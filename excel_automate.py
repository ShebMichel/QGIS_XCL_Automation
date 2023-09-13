# This file is used to automate excel
# import modules
import openpyxl, xlwings
import pandas as pd
from openpyxl import load_workbook
import os


def create_your_excel_file(filename,sheetname):
    '''Create, open, save, and close Excel files through Python scripts.
       # filename::  filename with its paths from the cwd()
       # sheetname:: excel sheet to manipulate data
    '''
    wb = openpyxl.Workbook()
    # Save the workbook
    wb.save(filename+'.xlsx')
    # Create sheetname
    wb.create_sheet(sheetname)
    wb.close()
    return print(filename+ ' saved and Open')


def sheet_manager(filename,sheetname,df):#col,col_header):
    ''' Add data table into excel and manipulate cell data.
        # filename::  filename with its paths from the cwd()
        # sheetname:: excel sheet to manipulate data
        # df: pandas dataframe
    '''
    gcw = os.getcwd()
    file=str(gcw)+'/'+str(filename)+'.xlsx'
   
    df.to_excel(file, sheet_name=str(sheetname))

    #writer.save()
    check_data=pd.read_excel(file, index_col=0) 
    print('below is the data check:')
    print(check_data)
    return

filename= str(input('Enter your file name: '))
sheetname   = str(input('Enter your sheetname: '))
df  = pd.DataFrame({'Data': [10, 20, 30, 20, 15, 30, 45]})
_          = create_your_excel_file(filename,sheetname)
sheet_manager(filename,sheetname, df)
